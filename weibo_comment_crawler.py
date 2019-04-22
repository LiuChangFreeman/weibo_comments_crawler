# -*- coding: utf-8 -*-
from __future__ import print_function
import datetime
import requests
import os
import io
import json
from bs4 import BeautifulSoup
import re
import openpyxl
dir="weibo_comment_3"
id="4360349987394836"
cookies=["ALF=1558189512; M_WEIBOCN_PARAMS=oid%3D4137015148682546%26luicode%3D20000061%26lfid%3D4137015148682546%26uicode%3D20000061%26fid%3D4137015148682546; SUHB=0IIFDMVgx6Svuh; _T_WM=cd29add9a894c8c0cdd2d319a3d487b6; SUBP=0033WrSXqPxfM725Ws9jqgMF55529P9D9W5-Ucj6vZT848KkzyOqM4Zi5JpX5KzhUgL.Foq0Soz4SK5ReK22dJLoI7RLxKML1K-L1h-ES0.RSntt; SUB=_2A25xvPiYDeRhGeBN7VAY9S7Eyj2IHXVTXpjQrDV6PUNbktAKLVXCkW1NRF0BTwXxofisG8Y-B7ZkVx7g0rfDfrrh; MLOGIN=1; WEIBOCN_FROM=1110006030; XSRF-TOKEN=0e1142","MLOGIN=1; _T_WM=9dffb4c6805017cbc1eedb8e65603a2b; M_WEIBOCN_PARAMS=oid%3D4137015148682546%26luicode%3D20000061%26lfid%3D4137015148682546%26uicode%3D20000061%26fid%3D4137015148682546; SUB=_2A25xvUorDeRhGeFO7FUS8ibKwj6IHXVTXlZjrDV6PUJbktANLRDYkW1NQVkEWk8Y37pyUZCexf2AM4Lrk_x-SnTm; SUBP=0033WrSXqPxfM725Ws9jqgMF55529P9D9WW8uO5.Fg6opimpfGig0eCz5JpX5KzhUgL.FoM7S0M0eonc1Kz2dJLoI05LxKMLB.2LBoMLxKML1h-LBKBLxKnLB.-LBKeLxK.LBKzL1K5LxKnLB--LB.q71KMt; SUHB=0wnlDJxDUyIjSo; XSRF-TOKEN=e9ec5e; WEIBOCN_FROM=1110106030","_T_WM=48934f72b1b1d783c341e88b0cd46874; WEIBOCN_FROM=1110106030; ALF=1558234525; MLOGIN=1; SCF=ArC_79Ko5gD4SagZ-jos_L37KCgFxX8XRZnm_UO0s5_8YpPPho_gMuUuV6j7hhRptsJ-9VhK0c76zqp7bgCDBM0.; SUB=_2A25xvUj1DeRhGeFO7FUS9ijLwzWIHXVTXmi9rDV6PUJbktAKLW3kkW1NQVkZ7Vm8_9ep_hCn9FoSHO-RqqdQ8bYp; SUBP=0033WrSXqPxfM725Ws9jqgMF55529P9D9W5fKeGTYSbh6REkMuNIQSAh5JpX5K-hUgL.FoM7S0M0SoqN1h.2dJLoI05LxKnLB.2LB.2LxK-L122LBK5LxKML1-zLB--LxKqLBK2LBoqLxK.L1h-LB-2N1Knt; SUHB=02M_jHJE3M287o; SSOLoginState=1555642533; M_WEIBOCN_PARAMS=oid%3D4137015148682546%26luicode%3D20000061%26lfid%3D4137015148682546%26uicode%3D20000061%26fid%3D4137015148682546; XSRF-TOKEN=eb647b","_T_WM=01fc1eb2aee1966df9d72834ffbbba86; ALF=1558198236; SUBP=0033WrSXqPxfM725Ws9jqgMF55529P9D9W5_2X18pk_MFGS0Zm3p_Hyu5JpX5K-hUgL.FoM7S0M41h.N1K22dJLoI05LxK-L1KqL1-eLxKnLBK2L1KBLxK-L1-zL1-zLxKqL12BLBKeLxKnLBo5L12qpeh2t; SCF=AiH-1-5jXI6iy2YKJoh8w0HRbgRxov8XQR2BkPc8Q5MipfzsreZUm_ah_62-Kc0-74Q-O6nUvtaiebyfVdNYifI.; SUB=_2A25xvNsyDeRhGeFO7FUY-CfLwj2IHXVTXuV6rDV6PUJbktANLVfGkW1NQVkZAn7Vp7Xu5cIxSQbOZ6-IjDBbyoCc; SUHB=0rrrYjmwRfZ-hx; MLOGIN=1; M_WEIBOCN_PARAMS=oid%3D4137015148682546%26luicode%3D20000061%26lfid%3D4137015148682546%26uicode%3D20000061%26fid%3D4137015148682546; XSRF-TOKEN=d2f6cc; WEIBOCN_FROM=1110006030"]
format="%a %b %d %H:%M:%S +0800 %Y"
session=requests.session()
def get_time():
    return datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
def get_xsrf(cookie):
    token=re.findall("(?<=XSRF-TOKEN=).{6}",cookie)
    return token[0]
def crawler():
    import time,random
    cookie=random.choice(cookies)
    filenames=[int(item.replace(".json","")) for item in os.listdir("{}/attachments".format(dir))]
    if len(filenames)==0:
        count=1
        max_id = ""
        max_id_type = 0
    else:
        max_filename=max(filenames)
        count = max_filename+1
        data = json.loads(io.open(os.path.join(dir, "attachments", "{}.json".format(max_filename)), encoding="utf-8").read())
        max_id = data["data"]["max_id"]
        max_id_type = data["data"]["max_id_type"]
    while True:
        print(count)
        if count==1:
            url="https://m.weibo.cn/comments/hotflow?id={}&mid={}&max_id_type=0".format(id, id)
        else:
            url="https://m.weibo.cn/comments/hotflow?id={}&mid={}&max_id={}&max_id_type={}".format(id, id,max_id,max_id_type)
        headers = {
            "Accept": "application/json, text/plain, */*",
            "MWeibo-Pwa": "1",
            "Referer": "https://m.weibo.cn/detail/4360349987394836",
            "User-Agent": "Mozilla/5.0 (Linux; Android 6.0; Nexus 5 Build/MRA58N) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/70.0.3538.25 Mobile Safari/537.36",
            "X-Requested-With": "XMLHttpRequest",
            "X-XSRF-TOKEN": get_xsrf(cookie),
            "Cookie": cookie
        }
        response = session.get(url, headers=headers)
        try:
            result = response.json()
            max_id=result["data"]["max_id"]
            max_id_type = result["data"]["max_id_type"]
        except:
            print(u"错误-{}".format(get_time()))
            time.sleep(180)
            while True:
                cookie_temp=random.choice(cookies)
                if cookie!=cookie_temp:
                    cookie=cookie_temp
                    break
            continue
        if max_id==0:
            break
        with open("{}/attachments/{}.json".format(dir,count),"w") as fd:
            fd.write(response.text)
        count+=1
        time.sleep(1)
def extract():
    def transfer_time(time):
        return datetime.datetime.strptime(time, format).strftime('%Y-%m-%d %H:%M:%S')
    total=[]
    for i in range(1,8891):
        filename="{}.json".format(i)
    # for filename in os.listdir("{}/attachments".format(dir)):
        print(filename)
        data=json.loads(io.open(os.path.join(dir,"attachments",filename),encoding="utf-8").read())
        for item in data["data"]["data"]:
            text=item["text"]
            blocks = re.findall("<span class=\"url-icon\">.+?</span>", text)
            for block in blocks:
                soup = BeautifulSoup(block, "html.parser")
                image = soup.find("img")
                replaced = image.attrs["alt"]
                text = text.replace(block, replaced)
            if "<" in text:
                soup = BeautifulSoup(text, "html.parser")
                text= soup.text
            date=transfer_time(item["created_at"])
            temp={
                "text":text,"date":date
            }
            total.append(temp)
    text = json.dumps(total, ensure_ascii=False, indent=4)
    if type(text)==str:
        text=text.decode("utf-8")
    with io.open(os.path.join(dir, "result.json"), "w", encoding="utf-8") as fd:
        fd.write(text)
def to_excel():
    data = json.loads(io.open(os.path.join(dir, "result.json"), encoding="utf-8").read())
    wb = openpyxl.Workbook()
    sheet = wb.create_sheet("data")
    count=1
    for item in data:
        print(count)
        sheet.cell(row=count, column=1).value = item["text"]
        sheet.cell(row=count, column=2).value = item["date"]
        count+=1
    wb.save(u"置顶微博爬虫.xlsx")
def main():
    init()
    crawler()
    # extract()
    # to_excel()
def init():
    global result,crawlered,downloaded,record
    if not os.path.exists(dir):
        os.mkdir(dir)
    if not os.path.exists(os.path.join(dir,"attachments")):
        os.mkdir(os.path.join(dir,"attachments"))
    if not os.path.exists(os.path.join(dir,"backup")):
        os.mkdir(os.path.join(dir,"backup"))
if __name__=="__main__":
    main()